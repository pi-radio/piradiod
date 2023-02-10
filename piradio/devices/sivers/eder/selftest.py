import time

from piradio.devices.sivers.eder.registers import attach_registers, set_bits, clear_bits

def execute_self_test(port):
    pname = "A" if port == 0 else "B"

    if (eder.bf_rx_mbist_done != 0) or (eder.bf_tx_mbist_done != 0):
        raise RuntimeError(f"Port {pname}: Self test active at beginning")
    
    eder.bf_rx_mbist_2p_sel = port
    eder.bf_tx_mbist_2p_sel = port
    
    eder.bf_rx_mbist_en = 0xFFFF
    eder.bf_tx_mbist_en = 0xFFFF
    
    time.sleep(0.01)
    
    if (bf_rx_mbist_done != 0xffff) or (bf_tx_mbist_done != 0xffff):
        raise RuntimeError(f"Port {pname}: Self test timed out")
    
    if (bf_rx_mbist_result != 0) or (bf_tx_mbist_result != 0):
        raise RuntimeError(f"Port {pname}: Self test failed")
    
    eder.bf_rx_mbist_en = 0x0000
    eder.bf_tx_mbist_en = 0x0000
    
def self_test(eder, ports=0x3):
    PORT_A = 0x1
    PORT_B = 0x2

    if ports & PORT_A:
        execute_self_test(0)

    if ports & PORT_B:
        execute_self_test(1)

def agc_test(eder):
    """
        self.eder.fpga_clk(1)

        self.eder.regs.wr('agc_en', 0x15)
        self.eder.regs.wr('agc_timeout', 200)
        self.eder.regs.wr('agc_use_agc_ctrls', 0x3F)
        self.eder.regs.wr('agc_detector_mask', 0x1F1F)
        self.eder.regs.wr('agc_bf_rf_gain_lvl', 0x55443322)
        self.eder.regs.wr('agc_bb_gain_1db_lvl', 0x654321)

        #self.eder.regs.wr('gpio_agc_done_ctrl', 0x02)

        self.eder.evkplatform.drv.setagcrst(1)
        time.sleep(0.01)
        self.eder.evkplatform.drv.setagcrst(0)

        self.eder.evkplatform.drv.setagcstart(1)
        time.sleep(0.01)
        self.eder.evkplatform.drv.setagcstart(0)

        agc_status = self.eder.evkplatform.drv.getagcstate()
        while (agc_status & 0x80) == 0:
            agc_status = self.eder.evkplatform.drv.getagcstate()
        
        print (hex(self.eder.evkplatform.drv.getagcstate()))
    """
    pass

"""
                new_temp = self.eder.temp.run()-273
                self.eder.logger.log_info('Temp. ' + str(new_temp) + ' deg. C')
                self.eder.tx.alc.status()
                tx_alc_bfrf_gain = self.eder.regs.rd('tx_alc_bfrf_gain')
                time.sleep(1)
                spec.write("MKPK")
                #Read marker ampl.
                spec.write("MKA?")
                power = spec.read()
                writer.writerow([time_stamp, new_temp, power, tx_alc_bfrf_gain])
                #if (new_temp - temp > 4):
                #    low_th = low_th 
                #sa.write('CALC:MARK2:Y?')
                #spec_pow = sa.read()
                #print (spec_pow)
                #print ('')
                time.sleep(0.1)
                self.eder.evkplatform.drv.settxrxsw(0)
                self.eder.logger.log_info('TX Off')


    def test_lo_calib(self, num_of_tries=10):
        num_of_failures = 0
        freqs = [58.32e9, 60.48e9, 62.64e9, 64.80e9, 66.96e9, 69.12e9]
        for i in range(0, num_of_tries):
            for freq in freqs:
                print ('Temperature: ' + str(self.eder.temp.run()-273) + ' deg. C')
                self.eder.run_tx(freq)
                if self.eder.tx.dco.run() == False:
                    num_of_failures = num_of_failures + 1
                    print ('Calibration Failure')
                    print ('    Temperature: ' + str(self.eder.temp.run()-273) + ' deg. C')
                    print ('    Freq: ' + str(freq) + ' Hz')
        print ('Number of failures: ' + str(num_of_failures) + ' of ' + str(num_of_tries*6))

    def set_gain_index(self, gain_index):
        self.eder.regs.wr('rx_gain_ctrl_bfrf', self.agc_gain_table[gain_index][0]<<4|self.agc_gain_table[gain_index][1])
        self.eder.regs.wr('rx_gain_ctrl_bb1', self.agc_gain_table[gain_index][2]<<4|self.agc_gain_table[gain_index][2])
        self.eder.regs.wr('rx_gain_ctrl_bb2', self.agc_gain_table[gain_index][3]<<4|self.agc_gain_table[gain_index][3])
        self.eder.regs.wr('rx_gain_ctrl_bb3', self.agc_gain_table[gain_index][4]<<4|self.agc_gain_table[gain_index][4])

    def test_dco_screening(self):
        import time

        self.eder.run_rx()
        self.eder.regs.wr('trx_rx_on', 0x1fffff)
        self.set_gain_index(40)
        time.sleep(1)
        if self.eder.regs.device_info.get_attrib('chip_type') == 'Eder B MMF':
            self.eder.rx.drv_dco.run()
        self.eder.rx.dco.run()
        self.eder.regs.dump('rx')
        diff = self.eder.rx.dco.iq_meas.meas_vdiff(num_samples=16)
        print(round(1000 * self.eder.test._AdcToVolt(diff['idiff'])/(-2.845)))
        print(round(1000 * self.eder.test._AdcToVolt(diff['qdiff'])/(-2.845)))
        i_dco = []
        q_dco = []
        i_dco_beam = []
        q_dco_beam = []
        for beam in range(0,64):
            self.eder.rx.set_beam(beam)
            time.sleep(0.1)
            diff = self.eder.rx.dco.iq_meas.meas_vdiff(num_samples=16)
            i_dco.append(round(1000 * self.eder.test._AdcToVolt(diff['idiff'])/(-2.845)))
            q_dco.append(round(1000 * self.eder.test._AdcToVolt(diff['qdiff'])/(-2.845)))
            i_dco_beam.append(beam)
            q_dco_beam.append(beam)

        log_string = 'I RX DCO[mV] '+ ' '.join('{:3.0f}[{}]'.format(dco,dco_beam) for dco,dco_beam in zip(i_dco,i_dco_beam))
        print (log_string)

        log_string = 'Q RX DCO[mV] '+ ' '.join('{:3.0f}[{}]'.format(dco,dco_beam) for dco,dco_beam in zip(q_dco,q_dco_beam))
        print (log_string)

        self.eder.reset()

    def dco_beam_sweep_b_magn_phase(self, gain_index=None, file_name='dco_beam_sweep_b_magn_phase.xlsx', numsamples=16, freq=None, mag_max_y=None):
        from openpyxl import Workbook
        import time
        import math
        import numpy as np
        import matplotlib.pyplot as plt

        if gain_index != None:
            calib_gain_index = gain_index
            self.eder.regs.wr('rx_gain_ctrl_bfrf', self.agc_gain_table[calib_gain_index][0]<<4|self.agc_gain_table[calib_gain_index][1])
            self.eder.regs.wr('rx_gain_ctrl_bb1', self.agc_gain_table[calib_gain_index][2]<<4|self.agc_gain_table[calib_gain_index][2])
            self.eder.regs.wr('rx_gain_ctrl_bb2', self.agc_gain_table[calib_gain_index][3]<<4|self.agc_gain_table[calib_gain_index][3])
            self.eder.regs.wr('rx_gain_ctrl_bb3', self.agc_gain_table[calib_gain_index][4]<<4|self.agc_gain_table[calib_gain_index][4])
        else:
            self.eder.regs.wr('rx_gain_ctrl_bfrf', 0xff)
            self.eder.regs.wr('rx_gain_ctrl_bb1', 0xff)
            self.eder.regs.wr('rx_gain_ctrl_bb2', 0x33)
            self.eder.regs.wr('rx_gain_ctrl_bb3', 0x33)
            pass

        self.eder.regs.dump('rx')

        self.eder.regs.wr('trx_rx_on', 0x1fffff)
        time.sleep(10)
        if self.eder.regs.device_info.get_attrib('chip_type') == 'Eder B MMF':
            self.eder.rx.drv_dco.run()
        self.eder.rx.dco.run()

        wb = Workbook()
        ws = wb.active

        ws.append(["Beam", "Temp [deg. C]", "i_diff[ADC]", " q_diff[ADC]", "i_diff[V]", " q_diff[V]", "Ampl.", "Phase[deg.]"])

        ampl = []
        phase = []

        for beam in range(0,64):
            self.eder.rx.set_beam(beam)
            time.sleep(0.1)
            diff = self.eder.rx.dco.iq_meas.meas_vdiff(num_samples=numsamples)
            row = [beam, self.eder.temp.run()-273, diff['idiff'], diff['qdiff'], (self._AdcToVolt(diff['idiff'])/(-2.845))/2, (self._AdcToVolt(diff['qdiff'])/(-2.845))/2]
            row.append(math.sqrt(row[4]*1000000*row[4]+row[5]*1000000*row[5]))
            row.append(180*math.atan2(row[4],row[5])/math.pi)
            ws.append(row)
            ampl.append(row[6])
            phase.append(row[7])

        wb.save(file_name)

        self.plot_ampl_phase(file_name, title='Freq. '+str(freq), mag_max_y=30, save_plot=True)

    def plot_ampl_phase(self, file_name, title=None, worksheet=None, mag_max_y=None, save_plot=False):
        from openpyxl import Workbook
        from openpyxl import load_workbook
        import math
        import numpy as np
        import matplotlib.pyplot as plt

        wb = load_workbook(filename = file_name)
        if worksheet == None:
            ws = wb.active
        else:
            ws = wb[worksheet]

        ampl = []
        for col in ws['G']:
            ampl.append(col.value)
        ampl = ampl[1:len(ampl)]

        phase = []
        for col in ws['H']:
            phase.append(col.value)
        phase = phase[1:len(phase)]

        fig, axs = plt.subplots(2, figsize=[18,12])
        if title != None:
            fig.suptitle(str(title))
        plt.setp(axs[0], xticks=range(0,64))
        plt.setp(axs[1], xticks=range(0,64))
        axs[0].plot(range(0,64), ampl, 'tab:orange')
        axs[0].set_title('Magnitude')
        if mag_max_y == None:
            axs[0].set(xlabel='Beam', ylabel='Magnitude [mV]')
        else:
            axs[0].set(xlabel='Beam', ylabel='Magnitude [mV]', ylim=(0, mag_max_y))
        axs[1].plot(range(0,64), phase, 'tab:red')
        axs[1].set_title('Phase [deg.]')
        axs[1].set(xlabel='Beam', ylabel='Phase [deg.]', ylim=(-180, 180))

        plt.show(block=False)
        plt.pause(1)
        plt.savefig(file_name.replace('.xlsx', '.png'),format='png')

    def dco_beam_chan_sweep(self, prefix=None, gain_index=42, mag_max_y=30):
        channels = [58.32e9, 60.48e9, 62.64e9, 64.80e9, 66.96e9, 69.12e9]
        if prefix==None:
            base_file_name = ''
        else:
            base_file_name = str(prefix)+'_'
        for freq in channels:
            self.eder.run_rx(freq)
            self.dco_beam_sweep_b_magn_phase(gain_index=gain_index, file_name='dco_meas_data\\'+base_file_name+str(freq)+ '.xlsx', freq=freq, mag_max_y=mag_max_y)
            #self.dco_beam_sweep_b_magn_phase(gain_index=0, file_name='dco_meas_data\\'+base_file_name+str(freq)+ '.xlsx', freq=freq, mag_max_y=None)


    def temp_accuracy_check(self):
        import time
        from openpyxl import Workbook
        from openpyxl import load_workbook
        import pyvisa
        rm = pyvisa.ResourceManager()
        mm = rm.open_resource('GPIB0::22::INSTR')

        file_name = 'temp_meas.xlsx'
        #wb = load_workbook('temp_meas.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.append(["trx_tx_on", "Sensor[mV]", "PCB temp. [deg. C]", "Chip temp.[deg. C]"])

        self.eder.pll.init()
        self.eder.pll.set(60.48e9)
        self.eder.tx.setup(60.48e9, beam=31, trx_tx_on=0x1f0000)
        self.eder.regs.wr('trx_ctrl', 0x02)
        trx_tx_on=0x1f0000
        for i in range(0,16):
            print ('trx_tx_on: '+hex(trx_tx_on))
            for k in range(0,60):
                time.sleep(1)
                chip_temp = self.eder.temp.run('C')
                pcb_temp = -300
                while pcb_temp == -300:
                    pcb_temp = self.eder.eeprom.read_pcb_temp()
                mm.write('MEASure:VOLTage:DC?')
                try:
                    sensor_v = float(mm.read())*1000
                except:
                    sensor_v = 0
                print ('CHIP temperature: {:5.1f} Celcius'.format(chip_temp))
                print ('Sensor temperature: {:5.1f} Celcius'.format(sensor_v))
                ws.append([hex(trx_tx_on), sensor_v, pcb_temp, chip_temp])
                wb.save(file_name)
            trx_tx_on = trx_tx_on + (1 << i)
            self.eder.regs.wr('trx_tx_on', trx_tx_on)
"""
